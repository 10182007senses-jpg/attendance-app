import os, csv, io
from urllib.parse import quote
from datetime import datetime, date, time, timedelta
import secrets
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import pandas as pd
from fastapi import FastAPI, Depends, HTTPException, status, Cookie, Response
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi import Header
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session

from db import Session as DbSession, SessionLocal, init_db, User, AttendanceLog, Workday, UserMonthlyRequirement, MonthlyConfirmation, init_engine, Base, engine, now_jst_naive
from security import verify_pin, hash_pin
from seed_users import main as seed_main
from seed_users import seed_users


SESSION_RETENTION_DAYS = 30
DEFAULT_USER = "山田太郎"
SESSION_TTL_HOURS = 8       
IDLE_TIMEOUT_MINUTES = 1000  


ACTION_IN = "入室"
ACTION_OUT = "退室"

STATE_NOT_IN = "未入室"
STATE_IN_ROOM = "在室中"
STATE_UNKNOWN = "不明"
STATE_VALUES = {STATE_NOT_IN, STATE_IN_ROOM, STATE_UNKNOWN}
ROUND_MINUTES = 30
DEFAULT_REQUIRED_HOURS = 160

STATIC_DIR = "static"
INDEX_FILE = os.path.join(STATIC_DIR, "index.html")
ADMIN_FILE = os.path.join(STATIC_DIR, "admin.html")

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill("solid", fgColor="F2F2F2")

app = FastAPI()
bearer_scheme = HTTPBearer(auto_error=False)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

def get_session_id_from_cookie(
    session: str | None = Cookie(default=None)
) -> str | None:
  return session

def df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
  buf = io.BytesIO()

  with pd.ExcelWriter(buf, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name=sheet_name)

    ws = writer.sheets[sheet_name]
    ws.freeze_panes = "A2"

    for col in ws.columns:
      max_len = 0
      col_letter = col[0].column_letter
      for cell in col:
        v = "" if cell.value is None else str(cell.value)
        if len(v) > max_len:
          max_len = len(v)
      
      ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

  return buf.getvalue()

def _set_col_width(ws, widths: dict[int, float]):
  for col_idx, w in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = w

def _set_outer_border(ws, cell_range: str):
  rows = list(ws[cell_range])
  if not rows:
    return
  top = rows[0]
  bottom = rows[-1]
  for c in top:
    c.border = Border(
      left=c.border.left,
      right=c.border.right,
      top=Side(style="medium", color="666666"),
      bottom=c.border.bottom,
    )
  for c in bottom:
    c.border = Border(
      left=c.border.left,
      right=c.border.right,
      top=c.border.top,
      bottom=Side(style="medium", color="666666"),
    )
  for row in rows:
    left = row[0]
    right = row[-1]
    left.border = Border(
      left=Side(style="medium", color="666666"),
      right=left.border.right,
      top=left.border.top,
      bottom=left.border.bottom,
    )
    right.border = Border(
      left=right.border.left,
      right=Side(style="medium", color="666666"),
      top=right.border.top,
      bottom=right.border.bottom,
    )

def _style_range(ws, cell_range: str, bold=False, fill=False, center=False):
  for row in ws[cell_range]:
    for c in row:
      if bold:
        c.font = Font(bold=True)
      if fill:
        c.fill = FILL_HEADER
      if center:
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
      else:
        c.alignment = Alignment(vertical="center", wrap_text=True)
      c.border = BORDER

def build_user_month_xlsx(
    *,
    month: str,
    user: str,
    scheduled_minutes: int,
    confirmed_name: str | None = None,
    confirmed_at: datetime | None = None,
    details: list[dict],
    raw_logs: list[dict],
) -> bytes:
  wb = Workbook()

  ws = wb.active
  ws.title = "月次サマリー"
  ws.freeze_panes = "A9"
  ws.print_title_rows = f"{8}:{8}"
  ws.page_setup.paperSize = ws.PAPERSIZE_A4
  ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
  ws.page_setup.fitToWidth = 1
  ws.page_setup.fitToHeight = 0
  ws.page_margins.left = 0.3
  ws.page_margins.right = 0.3
  ws.page_margins.top = 0.5
  ws.page_margins.bottom = 0.5

  ws["A1"] = "勤怠確認票（月次）"
  ws["A1"].font = Font(bold=True, size=14)
  ws.merge_cells("A1:D1")
  ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
  ws["E1"] = "Quesera Grace 株式会社"
  ws["E1"].font = Font(bold=True, size=12)
  ws.merge_cells("E1:G1")
  ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
  ws.row_dimensions[1].height = 26

  # Header block
  ws["A3"] = "対象月"
  ws["C3"] = month
  ws["E3"] = "氏名"
  ws["F3"] = user
  ws["A4"] = "所定労働時間"
  ws["C4"] = f"{scheduled_minutes//60}時間{scheduled_minutes%60}分"
  ws.merge_cells("C3:D3")
  ws.merge_cells("C4:D4")
  ws.merge_cells("F3:G3")
  _style_range(ws, "A3:G4", bold=False, fill=True, center=False)
  _style_range(ws, "A3:A4", bold=True, fill=False, center=False)
  _style_range(ws, "E3:E3", bold=True, fill=False, center=False)
  _set_outer_border(ws, "A3:G4")
  label_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
  ws["A3"].alignment = label_align
  ws["A4"].alignment = label_align
  ws["E3"].alignment = label_align
  ws.row_dimensions[3].height = 20
  ws.row_dimensions[4].height = 20

  header_row = 8
  headers = ["日付", "入室(実打刻)", "退室(実打刻)", "入室(丸め後)", "退室(丸め後)", "実働時間", "備考"]
  for i, h in enumerate(headers, start=1):
    ws.cell(row=header_row, column=i, value=h)

  _set_col_width(ws, {
    1: 12, 2: 16, 3: 16, 4: 16, 5: 16, 6: 14, 7: 28
  })

  _style_range(ws, f"A{header_row}:G{header_row}", bold=True, fill=True, center=True)
  _set_outer_border(ws, f"A{header_row}:G{header_row}")
  ws.row_dimensions[header_row].height = 20
  for c in ws[f"A{header_row}:G{header_row}"][0]:
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

  total_net_min = 0
  actual_days = 0
  bad_days = 0

  row = header_row + 1
  for d in details:
    actions = d.get("action") or []
    in_times = [a["time"] for a in actions if a.get("action") == "入室"]
    out_times = [a["time"] for a in actions if a.get("action") == "退室"]
    raw_in = in_times[0] if in_times else ""
    raw_out = out_times[-1] if out_times else ""

    ok = bool(d.get("ok"))
    err = d.get("system_note") or d.get("error") or ""
    remark = d.get("remark") or merge_day_notes(d.get("employee_note"), err)
    if actions:
      actual_days += 1
    if not ok:
      bad_days += 1

    net_str = d.get("net") or ""
    if ok and isinstance(net_str, str) and ("時間" in net_str and "分" in net_str):
      try:
        h = int(net_str.split("時間")[0])
        m = int(net_str.split("時間")[1].split("分")[0])
        total_net_min += h * 60 + m
      except Exception:
        pass
    
    ws.cell(row=row, column=1, value=d.get("date"))
    ws.cell(row=row, column=2, value=raw_in)
    ws.cell(row=row, column=3, value=raw_out)
    ws.cell(row=row, column=4, value=(d.get("start") or ""))
    ws.cell(row=row, column=5, value=(d.get("end") or ""))
    ws.cell(row=row, column=6, value=(d.get("net") or ""))
    ws.cell(row=row, column=7, value=remark)
    
    _style_range(ws, f"A{row}:G{row}", center=False)
    ws.row_dimensions[row].height = 18
    note = str(remark) if remark else ""
    note_len = len(note)
    ws[f"G{row}"].alignment = Alignment(vertical="top", wrap_text=True, shrink_to_fit=(note_len > 80))
    if note_len >= 20:
      ws.row_dimensions[row].height = 32
    if not ok and err:
      for c in ws[f"A{row}:G{row}"][0]:
        c.fill = PatternFill("solid", fgColor="FCE4D6")
        if c.column == 7:
          c.font = Font(color="9C0006")
    row += 1

  data_start_row = header_row + 1
  data_end_row = row - 1
  
  sum_row = row + 1
  ws.merge_cells(f"A{sum_row}:D{sum_row}")
  ws[f"A{sum_row}"] = "月次集計"
  _style_range(ws, f"A{sum_row}:D{sum_row}", bold=True, fill=True, center=True)

  ws[f"A{sum_row+1}"] = "実出勤日"
  ws[f"D{sum_row+1}"] = actual_days
  ws[f"A{sum_row+2}"] = "不整合日数"
  ws[f"D{sum_row+2}"] = bad_days
  ws[f"A{sum_row+3}"] = "実働合計(丸め後)"
  ws[f"D{sum_row+3}"] = f"{total_net_min//60}時間{total_net_min%60}分"
  ws[f"A{sum_row+4}"] = "所定労働時間(月)"
  ws[f"D{sum_row+4}"] = f"{scheduled_minutes//60}時間{scheduled_minutes%60}分"
  for r in range(sum_row + 1, sum_row + 5):
    ws.merge_cells(f"A{r}:C{r}")
  _style_range(ws, f"A{sum_row+1}:D{sum_row+4}", bold=False, fill=False, center=False)
  for r in range(sum_row + 1, sum_row + 5):
    ws[f"A{r}"].alignment = label_align
    ws[f"D{r}"].alignment = Alignment(horizontal="right", vertical="center")
  _set_outer_border(ws, f"A{sum_row}:D{sum_row+4}")

  sign_row = sum_row + 6
  ws.merge_cells(f"A{sign_row}:C{sign_row}")
  ws.merge_cells(f"D{sign_row}:E{sign_row}")
  ws.merge_cells(f"A{sign_row+2}:C{sign_row+2}")
  ws.merge_cells(f"D{sign_row+2}:E{sign_row+2}")
  ws[f"A{sign_row}"] = "本人署名："
  ws[f"D{sign_row}"] = confirmed_name or " "
  ws[f"F{sign_row}"] = "日付："
  ws[f"G{sign_row}"] = confirmed_at.strftime("%Y/%m/%d") if confirmed_at else "____/____/____"

  ws[f"A{sign_row+2}"] = "管理者確認："
  ws[f"D{sign_row+2}"] = " "
  ws[f"F{sign_row+2}"] = "日付："
  ws[f"G{sign_row+2}"] = "____/____/____"
  _style_range(ws, f"A{sign_row}:G{sign_row+2}", bold=False, fill=False, center=False)
  ws[f"A{sign_row}"].alignment = label_align
  ws[f"A{sign_row+2}"].alignment = label_align
  _set_outer_border(ws, f"A{sign_row}:G{sign_row+2}")

  note_col = None
  for c in range(1, 30):
    if ws.cell(row=header_row, column=c).value == "備考":
      note_col = c
      break
  if note_col is not None and data_end_row >= data_start_row:
    note_letter = get_column_letter(note_col)
    base_width = ws.column_dimensions[note_letter].width
    ws.column_dimensions[note_letter].width = (base_width + 8) if base_width else 36
    for c in ws[header_row]:
      a = c.alignment or Alignment()
      c.alignment = a.copy(wrap_text=False)
    for r in range(data_start_row, data_end_row + 1):
      note_cell = ws.cell(r, note_col)
      a = note_cell.alignment or Alignment()
      note_cell.alignment = a.copy(wrap_text=True, vertical="top")
      note_len = len(str(note_cell.value or ""))
      if note_len >= 20:
        ws.row_dimensions[r].height = 32

  ws2 = wb.create_sheet("生ログ")
  ws2.freeze_panes = "A2"
  raw_headers = ["ts", "user", "action", "lat", "lon", "source"]
  for i, h in enumerate(raw_headers, start=1):
    ws2.cell(row=1, column=i, value=h)
  _set_col_width(ws2, {1: 20, 2: 14, 3: 10, 4: 12, 5: 12, 6: 12})
  _style_range(ws2, "A1:F1", bold=True, fill=True, center=True)

  rr = 2
  for x in raw_logs:
    ws2.cell(rr, 1, x.get("ts", ""))
    ws2.cell(rr, 2, x.get("user", ""))
    ws2.cell(rr, 3, x.get("action", ""))
    ws2.cell(rr, 4, x.get("lat", ""))
    ws2.cell(rr, 5, x.get("lon", ""))
    ws2.cell(rr, 6, x.get("source", ""))
    _style_range(ws2, f"A{rr}:F{rr}", center=False)
    rr += 1

  buf = io.BytesIO()
  wb.save(buf)
  return buf.getvalue()
  
def cleanup_sessions(db: Session) -> int:
  now = now_jst_naive()
  cutoff = now - timedelta(days=SESSION_RETENTION_DAYS)

  q = (
    db.query(DbSession)
    .filter(
      or_(
        DbSession.expires_at < now,
        DbSession.revoked == True,
      ),
      DbSession.created_at < cutoff
    )
  )

  deleted = q.delete(synchronize_session=False)
  db.commit()
  return deleted

@app.on_event("startup")
def startup():
    init_db()

    with SessionLocal() as db:
        cleanup_sessions(db)

        if db.query(User).count() == 0:
            seed_users(db)
            
@app.get("/admin")
def admin_page():
  if os.path.exists(ADMIN_FILE):
    return FileResponse(ADMIN_FILE)
  raise HTTPException(status_code=404, detail="admin.html not found")

def get_db():
  db = SessionLocal()
  try:
    yield db
  finally:
    db.close()

ACTIONS = {ACTION_IN, ACTION_OUT}


def _get_user_id(db: Session, user_name: str) -> int | None:
  u = db.query(User).filter(User.name == user_name, User.is_active == True).one_or_none()
  return u.id if u else None

def month_range(month: str) -> tuple[datetime, datetime]:
  y, m = month.split("-")
  year = int(y)
  mon = int(m)
  start = datetime(year, mon, 1)
  
  if mon == 12:
    end = datetime(year+1, 1, 1)
  else:
    end = datetime(year, mon+1, 1)
  return start, end

def combine_jst_naive_datetime(day_value: str, time_value: str) -> datetime:
  try:
    return datetime.strptime(f"{day_value} {time_value}", "%Y-%m-%d %H:%M")
  except ValueError as exc:
    raise HTTPException(
      status_code=status.HTTP_400_BAD_REQUEST,
      detail="日付または時刻の形式が不正です",
    ) from exc

def parse_month_str(month_str: str) -> tuple[int, int]:
  try:
    year_str, month_part = month_str.split("-", 1)
    return int(year_str), int(month_part)
  except ValueError as exc:
    raise HTTPException(status_code=400, detail="month は YYYY-MM 形式で指定してください") from exc

def format_month_str(year: int, month: int) -> str:
  return f"{year}-{str(month).zfill(2)}"

def get_previous_month_str(base_dt: datetime | None = None) -> str:
  now = base_dt or now_jst_naive()
  first_day_this_month = datetime(now.year, now.month, 1)
  previous_month_last_day = first_day_this_month - timedelta(days=1)
  return format_month_str(previous_month_last_day.year, previous_month_last_day.month)

def get_month_confirmation(db: Session, user_id: int, month_str: str) -> MonthlyConfirmation | None:
  year, month = parse_month_str(month_str)
  return (
    db.query(MonthlyConfirmation)
    .filter(
      MonthlyConfirmation.user_id == user_id,
      MonthlyConfirmation.year == year,
      MonthlyConfirmation.month == month,
    )
    .one_or_none()
  )

def invalidate_month_confirmation(db: Session, user_id: int, target_dt: datetime) -> None:
  (
    db.query(MonthlyConfirmation)
    .filter(
      MonthlyConfirmation.user_id == user_id,
      MonthlyConfirmation.year == target_dt.year,
      MonthlyConfirmation.month == target_dt.month,
    )
    .delete(synchronize_session=False)
  )

def _today_range() -> tuple[datetime, datetime]:
  today = now_jst_naive().date()
  start = datetime.combine(today, time.min)
  end = start + timedelta(days=1)
  return start, end

def _ensure_workday(db: Session, user_id: int, d:date) -> Workday:
  wd = db.query(Workday).filter(Workday.user_id == user_id, Workday.date == d).one_or_none()
  if wd is None:
    wd = Workday(user_id=user_id, date=d, status="open", created_at=now_jst_naive(), updated_at=now_jst_naive())
    db.add(wd)
    db.flush()
  return wd

def add_log_db(db: Session, action: str, user_name: str, lat: float | None = None, lon: float | None = None):
  if action not in ACTIONS:
    raise ValueError("unknown action")
  
  user_id = _get_user_id(db, user_name)
  if user_id is None:
    raise ValueError("unknown user")
  
  now_dt = now_jst_naive()

  log = AttendanceLog(user_id=user_id, action=action, ts=now_dt, lat=lat, lon=lon, source=None)
  db.add(log)

  wd = _ensure_workday(db, user_id, now_dt.date())

  if action == ACTION_IN:
    wd.status = "open"
  elif action == ACTION_OUT:
    wd.status = "closed"
  wd.updated_at = now_jst_naive()

  db.commit()

  return {
    "ユーザー": user_name,
    "アクション": action,
    "時刻": now_dt.strftime("%Y-%m-%d %H:%M:%S"),
    "緯度": lat,
    "経度": lon,
  }

def recalc_workday(db: Session, user_id: int, d: date) -> dict:
  start = datetime.combine(d, time.min)
  end = start + timedelta(days=1)
  logs = (
    db.query(AttendanceLog)
    .filter(
      AttendanceLog.user_id == user_id,
      AttendanceLog.ts >= start,
      AttendanceLog.ts < end,
      AttendanceLog.action.in_(ACTIONS),
    )
    .order_by(AttendanceLog.ts.asc())
    .all()
  )

  wd = db.query(Workday).filter(Workday.user_id == user_id, Workday.date == d).one_or_none()
  result = calc_day_from_logs(logs)

  if not logs:
    if wd is not None:
      if wd.employee_note:
        wd.status = "open"
        wd.updated_at = now_jst_naive()
      else:
        db.delete(wd)
      db.commit()
    return result

  if wd is None:
    wd = _ensure_workday(db, user_id, d)

  wd.status = "closed" if logs[-1].action == ACTION_OUT else "open"
  wd.updated_at = now_jst_naive()
  db.commit()
  return result

def normalize_admin_action(action_type: str) -> str:
  mapping = {
    "clock_in": ACTION_IN,
    "clock_out": ACTION_OUT,
    ACTION_IN: ACTION_IN,
    ACTION_OUT: ACTION_OUT,
  }
  action = mapping.get(action_type)
  if action is None:
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="type は clock_in または clock_out を指定してください")
  return action

def get_last_action_db(db: Session, user_name:str) -> str | None:
  user_id = _get_user_id(db, user_name)
  if user_id is None:
    return None
  
  last = (
    db.query(AttendanceLog)
    .filter(AttendanceLog.user_id == user_id, AttendanceLog.action.in_(ACTIONS))
    .order_by(AttendanceLog.ts.desc())
    .first()
  )
  # Returns None when the user has no logs or doesn't exist.
  return last.action if last else None

def get_today_logs_db(db: Session, user_name: str):
  user_id = _get_user_id(db, user_name)
  if user_id is None:
    return []
  
  start, end = _today_range()
  logs = (db.query(AttendanceLog)
          .filter(
            AttendanceLog.user_id == user_id,
            AttendanceLog.ts >= start,
            AttendanceLog.ts < end,
            AttendanceLog.action.in_(ACTIONS),
          )
          .order_by(AttendanceLog.ts.asc())
          .all()
          )
  
  return [
    {
      "ユーザー": user_name,
      "アクション": l.action,
      "時刻": l.ts.strftime("%Y-%m-%d %H:%M:%S"),
      "緯度": l.lat,
      "経度": l.lon,
    }
    for l in logs
  ]

def get_current_state_db(db: Session, user_name: str):
  user_id = _get_user_id(db, user_name)
  if user_id is None:
    return {"state": STATE_NOT_IN, "last_action": None, "lat": None, "lon": None, "time": None}
  
  last = (
    db.query(AttendanceLog)
    .filter(AttendanceLog.user_id == user_id, AttendanceLog.action.in_(ACTIONS))
    .order_by(AttendanceLog.ts.desc())
    .first()
  )

  if not last:
    return {"state": STATE_NOT_IN, "last_action": None, "lat": None, "lon": None, "time": None}
  
  last_action = last.action

  if last_action is None or last_action == ACTION_OUT:
    state = STATE_NOT_IN
  elif last_action == ACTION_IN:
    state = STATE_IN_ROOM
  else:
    state = STATE_UNKNOWN

  last_time = last.ts.strftime("%Y-%m-%d %H:%M:%S") if hasattr(last.ts, "strftime") else (str(last.ts) if last.ts is not None else None)
  return {
    "state" : state,
    "last_action": last_action,
    "lat": last.lat,
    "lon": last.lon,
    "time": last_time,
  }

def group_logs_by_workday_start(logs: list["AttendanceLog"]) -> dict[date, list["AttendanceLog"]]:
  by_day: dict[date, list[AttendanceLog]] = defaultdict(list)

  current: list[AttendanceLog] = []
  start_day: date | None = None
  in_shift = False

  for l in logs:
    if l.action == ACTION_IN:
      if in_shift and start_day is not None:
        by_day[start_day].extend(current)
      current = [l]
      start_day = l.ts.date()
      in_shift = True
      continue

    if in_shift:
      current.append(l)

    if l.action == ACTION_OUT:
      if start_day is not None:
        by_day[start_day].extend(current)
      current = []
      start_day = None
      in_shift = False

  if in_shift and start_day is not None:
    by_day[start_day].extend(current)

  return dict(by_day)

def calc_work_time_db(db: Session, user_name: str):
    user_id = _get_user_id(db, user_name)
    if user_id is None:
      return {"error": "ユーザーが見つかりません"}
    
    start_dt, end_dt = _today_range()
    daylogs = (
      db.query(AttendanceLog)
      .filter(
        AttendanceLog.user_id == user_id,
        AttendanceLog.ts >= start_dt,
        AttendanceLog.ts < end_dt,
        AttendanceLog.action.in_(ACTIONS),
      )
      .order_by(AttendanceLog.ts.asc())
      .all()
    )
  
    r = calc_day_from_logs(daylogs)
    if not r["ok"]:
      return {"error": r["error"] or "計算できません"}
    
    return {
      "date": str(start_dt.date()),
      "gross_work_time": sec_to_hm(r["gross_sec"]),
      "net_work_time": sec_to_hm(r["net_sec"]),
      # 必要なら丸め前後を表示
      "start": r["start"].strftime("%H:%M"),
      "end": r["end"].strftime("%H:%M"), 
    }
      
def calc_day_from_logs(day_logs: list[AttendanceLog]) -> dict:
    if not day_logs:
      return {"ok": False, "gross_sec": 0, "break_sec": 0, "net_sec": 0, "error": "ログなし", "start": None, "end": None}
    
    day_logs = sorted(day_logs, key=lambda x : x.ts)
  
    ins = [l.ts for l in day_logs if l.action == ACTION_IN]
    outs = [l.ts for l in day_logs if l.action == ACTION_OUT]
    if not ins or not outs:
      return {"ok": False, "gross_sec": 0, "break_sec": 0, "net_sec": 0, "error": "入室または退室が不足", "start": (min(ins) if ins else None), "end": (max(outs) if outs else None)}
    
    raw_start = min(ins)
    raw_end = max(outs)
  
    start = ceil_time(raw_start, ROUND_MINUTES)
    end = floor_time(raw_end, ROUND_MINUTES)
  
    if end < start:
      return {"ok": False, "gross_sec": 0, "break_sec": 0, "net_sec": 0, "error": "入退室時刻が不正", "start": start, "end": end}
    
    gross_sec = int((end-start).total_seconds())
    if gross_sec < 0:
      gross_sec = 0
    break_sec = 3600 if gross_sec >= 6 * 3600 else 0
    net_sec = max(gross_sec - break_sec, 0)
    return {"ok": True, "gross_sec": gross_sec, "break_sec": break_sec, "net_sec": net_sec, "error": None, "start": start, "end": end}

def sec_to_hm(sec: int) ->str:
  minutes = sec // 60
  h = minutes // 60
  m = minutes % 60
  return f"{h}時間{m}分"

def ceil_time(dt: datetime, minutes: int = ROUND_MINUTES) -> datetime:
  dt0 = dt.replace(second=0, microsecond=0)
  m = dt0.minute

  if (m % minutes) == 0:
    return dt0
  
  add = minutes - (m % minutes)
  return dt0 + timedelta(minutes=add)

def floor_time(dt: datetime, minutes: int = ROUND_MINUTES) -> datetime:
  dt0 = dt.replace(second=0, microsecond=0)
  m = dt0.minute
  if (m % minutes) == 0:
    return dt0 - timedelta(minutes=minutes)
  sub = (m % minutes)
  return dt0 - timedelta(minutes=sub)

def _get_bearer_token(authorization: str | None) -> str | None:
  if not authorization:
    return None
  elif not authorization.startswith("Bearer "):
    return None
  return authorization.removeprefix("Bearer ").strip() or None

def get_acurrent_user_row(creds: HTTPAuthorizationCredentials | None = Depends(bearer_scheme), db: Session = Depends(get_db)) -> User | None:
  if not creds:
    print("[auth] no credentials")
    return None
  
  session_id = creds.credentials

  now = now_jst_naive()
  s = db.query(DbSession).filter(DbSession.id == session_id).one_or_none()
  if not s:
    print("[auth] session not found:", session_id)
    return None
  if s.revoked:
    print("[auth] session revoked:", session_id)
    return None
  if s.expires_at <= now:
    print("[auth] session expired:", session_id, "expires_at=", s.expires_at, "now=", now)
    s.revoked = True
    db.commit()
    return None
  
  if s.last_seen_at and (now - s.last_seen_at) > timedelta(minutes=IDLE_TIMEOUT_MINUTES):
    print("[auth] idle timeout:", session_id, "last_seen_at=", s.last_seen_at, "now=", now)
    s.revoked = True
    db.commit()
    return None
  
  s.last_seen_at = now
  db.commit()

  u = db.query(User).filter(User.id == s.user_id, User.is_active == True).one_or_none()
  if not u:
    print("[auth] user not found or inactive:", s.user_id)
  return u

def get_current_user_row_from_cookie(
    session_id: str | None = Depends(get_session_id_from_cookie),
    db: Session = Depends(get_db),
) -> User | None:
  if not session_id:
    return None
  
  now = now_jst_naive()
  s = db.query(DbSession).filter(DbSession.id == session_id).one_or_none()
  if not s or s.revoked or s.expires_at <= now:
    if s and s.expires_at <= now:
      s.revoked = True
      db.commit()
    return None
    
  if s.last_seen_at and now - s.last_seen_at > timedelta(minutes=IDLE_TIMEOUT_MINUTES):
    s.revoked = True
    db.commit()
    return None

  s.last_seen_at = now
  db.commit()

  u = db.query(User).filter(User.id == s.user_id, User.is_active == True).one_or_none()
  return u


def require_admin(user: User | None = Depends(get_current_user_row_from_cookie)) -> User:
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")
  if user.role != "admin":
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="権限がありません")
  return user

class LoginRequest(BaseModel):
  user: str
  pin:str

class AdminMonthlyHoursSaveRequest(BaseModel):
  user_id: int
  year: int
  data: dict[str, float]

class AdminLogUpdateRequest(BaseModel):
  ts: datetime

class AdminLogCreateRequest(BaseModel):
  user_id: int
  type: str
  ts: datetime

class AdminFullDayLogCreateRequest(BaseModel):
  user_id: int
  date: str
  clock_in: str
  clock_out: str

class AdminFillMissingLogRequest(BaseModel):
  user_id: int
  date: str
  type: str
  time: str

class MonthConfirmRequest(BaseModel):
  month: str

class TodayNoteRequest(BaseModel):
  note: str

class AdminCreateUserRequest(BaseModel):
  name: str
  pin: str
  role: str = "user"

class AdminUpdateUserRequest(BaseModel):
  name: str
  role: str | None = None
  is_active: bool | None = None
  pin: str | None = None

def get_current_user(session_id: str | None = Depends(get_session_id_from_cookie) , db: Session = Depends(get_db)):
  if not session_id:
    return None
  
  now = now_jst_naive()

  s = (
    db.query(DbSession).filter(DbSession.id == session_id).one_or_none()
  )

  if not s:
    return None
  if s.revoked:
    return None
  if s.expires_at <= now:
    s.revoked = True
    db.commit()
    return None
  
  if s.last_seen_at and now - s.last_seen_at > timedelta(minutes=IDLE_TIMEOUT_MINUTES):
    s.revoked = True
    db.commit()
    return None
  
  s.last_seen_at = now
  db.commit()

  u = db.query(User).filter(User.id == s.user_id, User.is_active == True).one_or_none()
  return u.name if u else None


def get_month_logs_for_user(db: Session, user_id: int, month: str) -> list[AttendanceLog]:
  start, end = month_range(month)
  return (
    db.query(AttendanceLog)
    .filter(AttendanceLog.user_id == user_id, AttendanceLog.ts >= start, AttendanceLog.ts < end, AttendanceLog.action.in_(ACTIONS))
    .order_by(AttendanceLog.ts.asc())
    .all()
  )

def merge_day_notes(employee_note: str | None, system_note: str | None) -> str:
  notes = []
  if employee_note:
    notes.append(employee_note)
  if system_note:
    notes.append(system_note)
  return "\n".join(notes)

def build_month_detail_rows(db: Session, user_id: int, month: str, logs: list[AttendanceLog]) -> list[dict]:
  day_map = group_logs_by_workday_start(logs)
  start, end = month_range(month)
  workdays = (
    db.query(Workday)
    .filter(
      Workday.user_id == user_id,
      Workday.date >= start.date(),
      Workday.date < end.date(),
    )
    .all()
  )
  workday_map = {wd.date: wd for wd in workdays}
  details = []
  all_dates = sorted(set(day_map.keys()) | set(workday_map.keys()))
  for d in all_dates:
    day_logs = day_map.get(d, [])
    wd = workday_map.get(d)
    employee_note = wd.employee_note if wd else None
    if day_logs:
      r = calc_day_from_logs(day_logs)
    else:
      r = {"ok": True, "error": None, "start": None, "end": None, "gross_sec": 0, "net_sec": 0}
    remark = merge_day_notes(employee_note, r["error"])
    details.append({
      "date": str(d),
      "ok": r["ok"],
      "error": r["error"],
      "system_note": r["error"],
      "employee_note": employee_note,
      "remark": remark,
      "start": r["start"].strftime("%H:%M:%S") if r["start"] else None,
      "end": r["end"].strftime("%H:%M:%S") if r["end"] else None,
      "gross": sec_to_hm(r["gross_sec"]) if day_logs else "",
      "net": sec_to_hm(r["net_sec"]) if day_logs else "",
      "action": [
        {"id": l.id, "action": l.action, "time": l.ts.strftime("%H:%M:%S"), "lat": l.lat, "lon": l.lon}
        for l in sorted(day_logs, key=lambda x: x.ts)
      ],
    })
  return details

def calc_month_summary_from_logs(
    logs: list[AttendanceLog],
    month: str,
    required_hours: float
):
  day_map = group_logs_by_workday_start(logs)

  worked_days = 0
  ok_days = 0
  inconsistent_days = 0
  total_net_sec = 0

  for d, day_logs in day_map.items():
    if not str(d).startswith(month):
      continue

    worked_days += 1
    r = calc_day_from_logs(day_logs)

    if r["ok"]:
      ok_days += 1
      total_net_sec += r["net_sec"]
    else:
      inconsistent_days += 1

  required_sec = int(required_hours * 3600)
  remaining_sec = max(required_sec - total_net_sec, 0)

  return {
    "month": month,
    "required_hours": required_hours,
    "worked_days": worked_days,
    "ok_days": ok_days,
    "inconsistent_days": inconsistent_days,
    "worked_time": sec_to_hm(total_net_sec),
    "remaining_time": sec_to_hm(remaining_sec),
    "remaining_minutes": remaining_sec // 60,
  }

def get_required_hours(db: Session, user_id: int, month_str: str) -> float:
  year, month = parse_month_str(month_str)

  req = db.query(UserMonthlyRequirement).filter(
    UserMonthlyRequirement.user_id == user_id,
    UserMonthlyRequirement.year == year,
    UserMonthlyRequirement.month == month,
  ).one_or_none()
  if req is not None:
    return req.required_hours

  u = db.query(User).filter(
    User.id == user_id,
    User.is_active == True 
    ).one_or_none()
  if u and u.required_hours is not None:
    return u.required_hours
  return DEFAULT_REQUIRED_HOURS

def get_user_monthly_hours_map(db: Session, user_id: int, year: int) -> dict[str, float]:
  u = db.query(User).filter(User.id == user_id, User.is_active == True).one_or_none()
  fallback = u.required_hours if u and u.required_hours is not None else DEFAULT_REQUIRED_HOURS
  data = {str(month): fallback for month in range(1, 13)}
  rows = (
    db.query(UserMonthlyRequirement)
    .filter(UserMonthlyRequirement.user_id == user_id, UserMonthlyRequirement.year == year)
    .all()
  )
  for row in rows:
    data[str(row.month)] = row.required_hours
  return data

@app.post("/login")
def login(req: LoginRequest, response: Response, db: Session = Depends(get_db)):
  user_name = req.user.strip()
  pin = req.pin.strip()

  u = db.query(User).filter(User.name == user_name, User.is_active == True).one_or_none()
  if not u:
    return {"ok": False, "error": "認証に失敗しました"}
  
  if not verify_pin(pin, u.pin_hash):
    return {"ok": False, "error": "認証に失敗しました"}
  
  session_id = secrets.token_hex(16)
  response.set_cookie(
    key="session",
    value=session_id,
    httponly=True,
    secure=bool(os.getenv("RENDER")),
    samesite="lax",
    max_age=SESSION_TTL_HOURS * 3600,
    path="/",
  )

  now = now_jst_naive()

  s = DbSession(
    id=session_id,
    user_id=u.id,
    created_at=now,
    expires_at=now + timedelta(hours=SESSION_TTL_HOURS),
    last_seen_at=now,
    revoked=False,
    device_label=None
  )
  db.add(s)
  db.commit()

  return {"ok": True, "user": u.name, "role": u.role}

@app.get("/admin/ping")
def admin_ping(admin: User = Depends(require_admin)):
  return {"ok": True, "user": admin.name}

@app.patch("/admin/log/{log_id}")
def admin_update_log(
  log_id: int,
  body: AdminLogUpdateRequest,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db),
):
  log = db.query(AttendanceLog).filter(AttendanceLog.id == log_id).one_or_none()
  if log is None:
    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="打刻が見つかりません")

  old_ts = log.ts
  old_date = log.ts.date()
  log.ts = body.ts
  invalidate_month_confirmation(db, log.user_id, old_ts)
  invalidate_month_confirmation(db, log.user_id, body.ts)
  db.commit()

  affected_dates = {old_date, body.ts.date()}
  results = {str(d): recalc_workday(db, log.user_id, d) for d in sorted(affected_dates)}
  return {"ok": True, "log_id": log.id, "user_id": log.user_id, "ts": log.ts.isoformat(), "results": results}

@app.post("/admin/log")
def admin_create_log(
  body: AdminLogCreateRequest,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db),
):
  user = db.query(User).filter(User.id == body.user_id, User.is_active == True).one_or_none()
  if user is None:
    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ユーザーが見つかりません")

  log = AttendanceLog(
    user_id=body.user_id,
    action=normalize_admin_action(body.type),
    ts=body.ts,
    lat=None,
    lon=None,
    source="admin",
  )
  db.add(log)
  invalidate_month_confirmation(db, body.user_id, body.ts)
  db.commit()
  db.refresh(log)

  result = recalc_workday(db, log.user_id, log.ts.date())
  return {"ok": True, "log_id": log.id, "user_id": log.user_id, "ts": log.ts.isoformat(), "result": result}

@app.post("/admin/logs/full-day")
def admin_create_full_day_logs(
  body: AdminFullDayLogCreateRequest,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db),
):
  user = db.query(User).filter(User.id == body.user_id, User.is_active == True).one_or_none()
  if user is None:
    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ユーザーが見つかりません")

  clock_in_dt = combine_jst_naive_datetime(body.date, body.clock_in)
  clock_out_dt = combine_jst_naive_datetime(body.date, body.clock_out)
  if clock_in_dt >= clock_out_dt:
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="入室時刻は退室時刻より前にしてください")

  day_start = datetime.combine(clock_in_dt.date(), time.min)
  day_end = day_start + timedelta(days=1)
  existing_log = (
    db.query(AttendanceLog)
    .filter(
      AttendanceLog.user_id == body.user_id,
      AttendanceLog.ts >= day_start,
      AttendanceLog.ts < day_end,
    )
    .first()
  )
  if existing_log is not None:
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="その日は既存ログがあります。個別編集を使ってください。")

  db.add_all([
    AttendanceLog(
      user_id=body.user_id,
      action=ACTION_IN,
      ts=clock_in_dt,
      lat=None,
      lon=None,
      source="admin",
    ),
    AttendanceLog(
      user_id=body.user_id,
      action=ACTION_OUT,
      ts=clock_out_dt,
      lat=None,
      lon=None,
      source="admin",
    ),
  ])
  invalidate_month_confirmation(db, body.user_id, clock_in_dt)
  db.commit()

  recalc_workday(db, body.user_id, clock_in_dt.date())
  return {"status": "ok"}

@app.post("/admin/logs/fill-missing")
def admin_fill_missing_log(
  body: AdminFillMissingLogRequest,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db),
):
  user = db.query(User).filter(User.id == body.user_id, User.is_active == True).one_or_none()
  if user is None:
    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ユーザーが見つかりません")

  action = normalize_admin_action(body.type)
  target_dt = combine_jst_naive_datetime(body.date, body.time)
  day_start = datetime.combine(target_dt.date(), time.min)
  day_end = day_start + timedelta(days=1)
  logs = (
    db.query(AttendanceLog)
    .filter(
      AttendanceLog.user_id == body.user_id,
      AttendanceLog.ts >= day_start,
      AttendanceLog.ts < day_end,
      AttendanceLog.action.in_(ACTIONS),
    )
    .order_by(AttendanceLog.ts.asc())
    .all()
  )
  if not logs:
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="その日はログがありません。1日分追加を使ってください。")

  if any(log.action == action for log in logs):
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{body.type}ログはすでに存在します。")

  if action == ACTION_IN:
    first_out = min((log.ts for log in logs if log.action == ACTION_OUT), default=None)
    if first_out is None:
      raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="補完できる退室ログがありません。")
    if target_dt >= first_out:
      raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="入室補完は既存の退室時刻より前にしてください。")
  elif action == ACTION_OUT:
    last_in = max((log.ts for log in logs if log.action == ACTION_IN), default=None)
    if last_in is None:
      raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="補完できる入室ログがありません。")
    if target_dt <= last_in:
      raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="退室補完は既存の入室時刻より後にしてください。")

  log = AttendanceLog(
    user_id=body.user_id,
    action=action,
    ts=target_dt,
    lat=None,
    lon=None,
    source="admin",
  )
  db.add(log)
  invalidate_month_confirmation(db, body.user_id, target_dt)
  db.commit()
  db.refresh(log)

  result = recalc_workday(db, log.user_id, log.ts.date())
  return {"ok": True, "log_id": log.id, "result": result}


@app.post("/logout")
def logout(response: Response, session_id: str | None = Depends(get_session_id_from_cookie), db: Session = Depends(get_db),):
  if not session_id:
    return {"ok": True}
  
  s = db.query(DbSession).filter(DbSession.id == session_id).one_or_none()
  if s:
    s.revoked = True
    db.commit()

  response.delete_cookie("session", path="/")
  return {"ok": True}

@app.get("/")
def root():
  if os.path.exists(INDEX_FILE):
    return FileResponse(INDEX_FILE)
  return {"message": "Attendance API is running"}

@app.get("/clock-in")
def clock_in(user: str = Depends(get_current_user), lat: float = None, lon: float = None, db: Session = Depends(get_db),):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")
  
  try:
    last = get_last_action_db(db, user)
    if last != ACTION_OUT and last is not None:
      return {"error": "すでに入室しています"}
      
    row = add_log_db(db, ACTION_IN, user, lat, lon)
    return {"status": "ok", "data": row}
  
  except Exception as e:
    print("clock_in error:", repr(e))
    return {"error":"入室の記録に失敗しました。"}

  
@app.get("/clock-out")
def clock_out(user: str = Depends(get_current_user), lat: float = None, lon: float = None, db: Session=Depends(get_db),):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")
  try:
    last = get_last_action_db(db, user)
    if last != ACTION_IN:
      return {"error": "入室していないため退室できません"}
    
    row = add_log_db(db, ACTION_OUT, user, lat, lon)
    return {"status": "ok", "data": row}
  except Exception as e:
    print("clock_out error:", repr(e))
    return {"error": "退室の記録に失敗しました"}

@app.get("/break-start")
def break_start(user: str = Depends(get_current_user), lat: float = None, lon: float = None, db: Session = Depends(get_db),):
  raise HTTPException(status_code=410, detail="この機能は現在利用できません")

@app.get("/break-end")
def break_end(user: str = Depends(get_current_user), lat: float = None, lon: float = None, db: Session = Depends(get_db),):
  raise HTTPException(status_code=410, detail="この機能は現在利用できません")

@app.get("/today-logs")
def today_logs(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")

  try:
    return {"logs": get_today_logs_db(db, user)}
  except Exception:
    return {"error": "ログの取得に失敗しました", "logs": []}

@app.get("/current-state")
def current_state(user: str = Depends(get_current_user), db: Session = Depends(get_db),):
  try:
    if user is None:
      return {"state": STATE_NOT_IN, "last_action": None, "lat": None, "lon": None, "time": None}
    return get_current_state_db(db, user)
  except Exception:
    return {"error": "状態の取得に失敗しました", "state": STATE_UNKNOWN}

@app.get("/work-time")
def work_time(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")

  try:
    return calc_work_time_db(db, user)
  except Exception:
    return {"error": "勤務時間の計算に失敗しました"}

@app.get("/today-note")
def today_note(user_row: User | None = Depends(get_current_user_row_from_cookie), db: Session = Depends(get_db)):
  if user_row is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")

  today = now_jst_naive().date()
  wd = (
    db.query(Workday)
    .filter(Workday.user_id == user_row.id, Workday.date == today)
    .one_or_none()
  )
  return {"status": "ok", "note": (wd.employee_note or "") if wd else ""}

@app.post("/today-note")
def save_today_note(
  body: TodayNoteRequest,
  user_row: User | None = Depends(get_current_user_row_from_cookie),
  db: Session = Depends(get_db),
):
  if user_row is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")

  note = body.note[:500]
  today = now_jst_naive().date()
  wd = _ensure_workday(db, user_row.id, today)
  wd.employee_note = note or None
  wd.updated_at = now_jst_naive()
  db.commit()
  return {"status": "ok"}

@app.get("/admin/month-summary")
def admin_month_summary(month, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
  start, end = month_range(month)

  users = db.query(User).filter(User.is_active == True).order_by(User.id.asc()).all()

  logs = (
    db.query(AttendanceLog)
    .filter(
      AttendanceLog.ts >= start,
      AttendanceLog.ts < (end + timedelta(days=1)),
      AttendanceLog.action.in_(ACTIONS),
    )
    .order_by(AttendanceLog.user_id.asc(), AttendanceLog.ts.asc())
    .all()
  )

  by_user_day: dict[int, dict[date, list[AttendanceLog]]] = {}
  for l in logs:
    by_user_day.setdefault(l.user_id, {}).setdefault(l.ts.date(), []).append(l)

  wds = (
    db.query(Workday)
    .filter(Workday.date >= start.date(), Workday.date < end.date())
    .all()
    )
  
  wd_map : dict[tuple[int, date]: Workday] = {(w.user_id, w.date): w for w in wds}

  rows = []

  for u in users:
    user_logs = [
      l for day_logs in by_user_day.get(u.id, {}).values() for l in day_logs
    ]
    day_map = group_logs_by_workday_start(user_logs)
    work_days = 0
    ok_days = 0
    inconsistent_day = 0
    open_days = 0

    total_gross = 0
    total_net = 0

    for d, day_logs in day_map.items():
      if not (start.date() <= d < end.date()):
        continue

      work_days += 1

      wd = wd_map.get((u.id, d))
      if wd and wd.status == "open":
        open_days += 1

      r = calc_day_from_logs(day_logs)
      if r["ok"]:
        ok_days += 1
        total_gross += r["gross_sec"]
        total_net += r["net_sec"]
      else:
        inconsistent_day += 1

    rows.append({
      "user_id": u.id,
      "user": u.name,
      "work_days_with_logs": work_days,
      "ok_days": ok_days,
      "inconsistent_days": inconsistent_day,
      "open_days": open_days,
      "total_net": sec_to_hm(total_net),
      "total_gross": sec_to_hm(total_gross),
    })

  return {
    "month": month,
    "rows": rows,
  }

@app.get("/admin/user-detail")
def admin_user_detail(
  month: str,
  user: str,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db),
):
  start, end = month_range(month)

  u = db.query(User).filter(User.name == user, User.is_active == True).one_or_none()
  if not u:
    return {"error": "対象のユーザーが見つかりません"}
  
  logs = (
    db.query(AttendanceLog)
    .filter(
      AttendanceLog.user_id == u.id,
      AttendanceLog.ts >= start,
      AttendanceLog.ts < end,
      AttendanceLog.action.in_(ACTIONS),
    )
    .order_by(AttendanceLog.ts.asc())
    .all()
  )

  return {
    "month": month,
    "user_id": u.id,
    "user": u.name,
    "details": build_month_detail_rows(db, u.id, month, logs),
  }

@app.get("/admin/user-detail.csv")
def admin_user_detail_csv(
  month: str,
  user: str,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db),
):
  data = admin_user_detail(month=month, user=user, admin=admin, db=db)

  if data.get("error"):
    return data
  
  output = io.StringIO()
  writer = csv.writer(output)

  writer.writerow([
    "month",
    "user",
    "date",
    "ok",
    "error",
    "start",
    "end",
    "gross",
    "net",
    "actions",
    "last_lat",
    "last_lon",
  ])

  for d in data["details"]:
    actions = ""
    if d.get("action"):
      actions = " / ".join([f'{a["action"]} {a["time"]} {a["lat"]} {a["lon"]}' for a in d["action"]])

    writer.writerow([
      data["month"],
      data["user"],
      d.get("date"),
      "1" if d.get("ok") else "0",
      d.get("error") or "",
      d.get("start") or "",
      d.get("end") or "",
      d.get("gross") or "",
      d.get("net") or "",
      actions,
    ])

  content = output.getvalue().encode("utf-8-sig")
  filename = f"user-detail_{month}_{user}.csv"
  filename_star = quote(filename)
  filename_ascii = filename.encode("ascii", "ignore").decode("ascii") or "user-detail.csv"

  return StreamingResponse(
    io.BytesIO(content),
    media_type="text/csv; charset=utf-8",
    headers={"Content-Disposition": f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{filename_star}'},
  )

@app.get("/admin/month-summary.csv")
def admin_month_summary_csv(
  month: str,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db)
):
  data = admin_month_summary(month=month, admin=admin, db=db)

  output = io.StringIO()
  writer = csv.writer(output)

  writer.writerow([
    "month",
    "user",
    "work_days_with_logs",
    "ok_days",
    "inconsistent_days",
    "open_days",
    "total_net",
    "total_gross",
  ])

  for row in data["rows"]:
    writer.writerow([
      data["month"],
      row["user"],
      row["work_days_with_logs"],
      row["ok_days"],
      row["inconsistent_days"],
      row["open_days"],
      row["total_net"],
      row["total_gross"],      
    ])

  content = output.getvalue().encode("utf-8-sig")
  filename = f"month-summary_{month}.csv"
  filename_star = quote(filename)
  filename_ascii = filename.encode("ascii", "ignore").decode("ascii") or "month-summary.csv"

  return StreamingResponse(
    io.BytesIO(content),
    media_type="text/csv; charset=utf-8",
    headers={"Content-Disposition": f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{filename_star}'},
  ) 


@app.get("/admin/raw-logs.csv")
def admin_raw_logs_csv(
  month: str,
  user: str | None = None,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db),
):
  start, end = month_range(month)

  q = (
    db.query(AttendanceLog, User)
    .join(User, User.id == AttendanceLog.user_id)
    .filter(
      AttendanceLog.ts >= start,
      AttendanceLog.ts < end,
      AttendanceLog.action.in_(ACTIONS),
    )
    .order_by(User.id.asc(), AttendanceLog.ts.asc())
  )

  if user:
    u = db.query(User).filter(User.name == user, User.is_active == True).one_or_none()
    if not u:
      raise HTTPException(status_code=404, detail="対象のユーザーが見つかりません")
    q = q.filter(AttendanceLog.user_id == u.id)
  
  rows = q.all()

  output = io.StringIO()
  writer = csv.writer(output)

  writer.writerow([
    "month",
    "user",
    "ts",
    "action",
    "lat",
    "lon",
    "source",
  ])

  for log, u in rows:
    writer.writerow([
      month,
      u.name,
      log.ts.strftime("%Y-%m-%d %H:%M:%S") if log.ts else "",
      log.action or "",
      "" if log.lat is None else log.lat,
      "" if log.lon is None else log.lon,
      "" if getattr(log, "souce", None) is None else log.source,
    ])

  content = output.getvalue().encode("utf-8-sig")
  filename =  f"raw-logs_{month}{('_' + user) if user else ''}.csv"
  filename_star = quote(filename)
  filename_ascii = filename.encode("ascii", "ignore").decode("ascii") or "raw-logs.csv"

  return StreamingResponse(
    io.BytesIO(content),
    media_type="text/csv; charset=utf-8",
    headers={"Content-Disposition": f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{filename_star}'},
  )

@app.get("/admin/user-monthly-hours")
def get_user_monthly_hours(
  user_id: int,
  year: int,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db)
):
  u = db.query(User).filter(User.id == user_id, User.is_active == True).one_or_none()
  if not u:
    raise HTTPException(status_code=404, detail="ユーザーが見つかりません")

  return {
    "ok": True,
    "status": "ok",
    "user_id": u.id,
    "year": year,
    "data": get_user_monthly_hours_map(db, u.id, year),
  }

@app.post("/admin/user-monthly-hours")
def save_user_monthly_hours(
  req: AdminMonthlyHoursSaveRequest,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db)
):
  u = db.query(User).filter(User.id == req.user_id, User.is_active == True).one_or_none()
  if not u:
    raise HTTPException(status_code=404, detail="ユーザーが見つかりません")

  for month in range(1, 13):
    raw_value = req.data.get(str(month), req.data.get(month))  # type: ignore[arg-type]
    hours = float(raw_value) if raw_value is not None else 0.0
    if hours < 0:
      raise HTTPException(status_code=400, detail="required_hoursは0以上")

    row = db.query(UserMonthlyRequirement).filter(
      UserMonthlyRequirement.user_id == u.id,
      UserMonthlyRequirement.year == req.year,
      UserMonthlyRequirement.month == month,
    ).one_or_none()
    if row is None:
      row = UserMonthlyRequirement(
        user_id=u.id,
        year=req.year,
        month=month,
        required_hours=hours,
      )
      db.add(row)
    else:
      row.required_hours = hours
      row.updated_at = now_jst_naive()

  db.commit()

  return {
    "ok": True,
    "status": "ok",
    "user_id": u.id,
    "year": req.year,
    "data": get_user_monthly_hours_map(db, u.id, req.year),
  }

@app.get("/admin/user-month.xlsx")
def admin_user_month_xlsx(
  month: str,
  user: str,
  admin: User = Depends(require_admin),
  db: Session = Depends(get_db)
):
  user_id = _get_user_id(db, user)
  if user_id is None:
    raise HTTPException(status_code=404, detail="対象のユーザーが見つかりません")
  month_required_hours = get_required_hours(db, user_id, month)
  month_required_minutes = int(month_required_hours * 60)
  confirmation = get_month_confirmation(db, user_id, month)

  detail = admin_user_detail(month, user, admin=admin, db=db)
  if detail.get("error"):
    return detail
  
  start, end = month_range(month)
  u = db.query(User).filter(User.name == user, User.is_active == True).one_or_none()
  if not u:
    raise HTTPException(status_code=404, detail="対象のユーザーが見つかりません")
  
  logs = (
    db.query(AttendanceLog)
    .filter(
      AttendanceLog.user_id == u.id,
      AttendanceLog.ts >= start,
      AttendanceLog.ts < end,
      AttendanceLog.action.in_(ACTIONS),
    )
    .order_by(AttendanceLog.ts.asc())
    .all()
  )

  raw_logs = []
  for l in logs:
    raw_logs.append({
      "ts": l.ts.strftime("%Y-%m-%d %H:%M:%S") if l.ts else "",
      "user": user,
      "action": l.action or "",
      "lat": "" if l.lat is None else l.lat,
      "lon": "" if l.lon is None else l.lon,
      "source": "" if getattr(l, "source", None) is None else l.source,      
    })
  
  content = build_user_month_xlsx(
    month=month,
    user=user,
    scheduled_minutes=month_required_minutes,
    confirmed_name=confirmation.confirmed_name if confirmation else None,
    confirmed_at=confirmation.confirmed_at if confirmation else None,
    details=detail["details"],
    raw_logs=raw_logs,
  )

  filename = f"attendance_{month}_{user}.xlsx"
  filename_star = quote(filename)
  filename_ascii = filename.encode("ascii", "ignore").decode("ascii") or "attendance.xlsx"

  return StreamingResponse(
    io.BytesIO(content),
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers={
      "Content-Disposition": f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{filename_star}'
    },
  )

@app.get("/admin/users")
def admin_list_users(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
  users = db.query(User).order_by(User.id.asc()).all()
  return {
    "ok": True,
    "users": [
      {"id": u.id, "name": u.name, "role": u.role, "is_active": u.is_active}
      for u in users
    ]
  }

@app.post("/admin/users")
def admin_create_user(req: AdminCreateUserRequest, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
  name = req.name.strip()
  pin = req.pin.strip()
  role = req.role.strip()

  if not name:
    raise HTTPException(status_code=400, detail="nameは必須です")
  if not pin:
    raise HTTPException(status_code=400, detail="pinは必須です")
  if role not in ("user", "admin"):
    raise HTTPException(status_code=400, detail="roleは user または admin")
  
  exists = db.query(User).filter(User.name == name).one_or_none()
  if exists:
    raise HTTPException(status_code=409, detail="同盟のユーザーが既に存在します")
  
  u = User(
    name=name,
    pin_hash=hash_pin(pin),
    role=role,
    is_active=True
  )
  db.add(u)
  db.commit()

  return {"ok": True, "user": {"id": u.id, "name": u.name, "role": u.role, "is_active": u.is_active}}

@app.patch("/admin/users")
def admin_update_user(req: AdminUpdateUserRequest, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
  name = req.name.strip()
  if not name:
    raise HTTPException(status_code=400, detail="nameは必須です")
  
  u = db.query(User).filter(User.name == name).one_or_none()
  if not u:
    raise HTTPException(status_code=404, detail="ユーザーが見つかりません")
  
  if req.role is not None:
    role = req.role.strip()
    if role not in ("user", "admin"):
      raise HTTPException(status_code=400, detail="roleは user または admin")
    u.role = role
  
  if req.is_active is not None:
    u.is_active = bool(req.is_active)

  if req.pin is not None:
    pin = req.pin.strip()
    if not pin:
      raise HTTPException(status_code=400, detail="pinが空です")
    u.pin_hash = hash_pin(pin)

  db.commit()

  return {"ok": True, "user": {"id": u.id, "name": u.name, "role": u.role, "is_active": u.is_active}}

@app.get("/me")
def me(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")
  u = db.query(User).filter(User.name == user, User.is_active == True).one_or_none()
  if not u:
    raise HTTPException(status_code=404, detail="ユーザーが見つかりません")
  return {"ok": True, "user": u.name, "role": u.role} 

@app.get("/me/previous-month-confirmation")
def my_previous_month_confirmation(
  user: str = Depends(get_current_user),
  db: Session = Depends(get_db),
):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")
  u = db.query(User).filter(User.name == user, User.is_active == True).one_or_none()
  if not u:
    raise HTTPException(status_code=404, detail="ユーザーが見つかりません")

  month = get_previous_month_str()
  confirmation = get_month_confirmation(db, u.id, month)
  if confirmation is None:
    return {"status": "ok", "month": month, "confirmed": False}

  return {
    "status": "ok",
    "month": month,
    "confirmed": True,
    "confirmed_at": confirmation.confirmed_at.isoformat(),
    "confirmed_name": confirmation.confirmed_name,
  }

@app.get("/me/previous-month-detail")
def my_previous_month_detail(
  user: str = Depends(get_current_user),
  db: Session = Depends(get_db),
):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")
  u = db.query(User).filter(User.name == user, User.is_active == True).one_or_none()
  if not u:
    raise HTTPException(status_code=404, detail="ユーザーが見つかりません")

  month = get_previous_month_str()
  logs = get_month_logs_for_user(db, u.id, month)
  required_hours = get_required_hours(db, u.id, month)
  summary = calc_month_summary_from_logs(logs, month, required_hours)
  confirmation = get_month_confirmation(db, u.id, month)
  return {
    "status": "ok",
    "month": month,
    "summary": summary,
    "details": build_month_detail_rows(db, u.id, month, logs),
    "confirmed": confirmation is not None,
    "confirmed_at": confirmation.confirmed_at.isoformat() if confirmation else None,
    "confirmed_name": confirmation.confirmed_name if confirmation else None,
  }

@app.post("/me/month-confirm")
def my_month_confirm(
  body: MonthConfirmRequest,
  user: str = Depends(get_current_user),
  db: Session = Depends(get_db),
):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")
  u = db.query(User).filter(User.name == user, User.is_active == True).one_or_none()
  if not u:
    raise HTTPException(status_code=404, detail="ユーザーが見つかりません")

  year, month = parse_month_str(body.month)
  confirmation = get_month_confirmation(db, u.id, body.month)
  if confirmation is None:
    confirmation = MonthlyConfirmation(
      user_id=u.id,
      year=year,
      month=month,
      confirmed_at=now_jst_naive(),
      confirmed_name=u.name,
    )
    db.add(confirmation)
    db.commit()

  return {"status": "ok"}

@app.get("/me/month-summary")
def my_month_summary(
  month: str | None = None,
  user: str = Depends(get_current_user),
  db: Session = Depends(get_db)
):
  if user is None:
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未ログイン")
  
  if not month:
    now = now_jst_naive()
    month = f"{now.year}-{str(now.month).zfill(2)}"

  user_id = _get_user_id(db, user)
  if user_id is None:
    raise HTTPException(status_code=404, detail="ユーザーが見つかりません")
  required_hours = get_required_hours(db, user_id, month)
  logs = get_month_logs_for_user(db, user_id, month)
  return calc_month_summary_from_logs(logs, month, required_hours)
